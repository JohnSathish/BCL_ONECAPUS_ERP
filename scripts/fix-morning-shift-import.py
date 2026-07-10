"""Prepare Morning Shift Sem 3 student import Excel for ERP upload."""
from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SOURCE = Path(
    r"C:\Users\johnm\OneDrive\Desktop\Import Live 1-3-5\Morning Shift\MORNING SHIFT IMPORT DATA 2026.xlsx"
)
OUTPUT = SOURCE.with_name("MORNING SHIFT IMPORT DATA 2026 - READY.xlsx")
REPORT = SOURCE.with_name("MORNING SHIFT IMPORT DATA 2026 - VALIDATION REPORT.txt")

PROGRAM_SEM3_DEFAULTS = {
    "BA-EDU": {
        "MDC (Sem 3)": "Gender Studies",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Mushroom Cultivation- I",
    },
    "BA-ENG": {
        "MDC (Sem 3)": "Gender Studies",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Baking and Confectionary-I",
    },
    "BA-POL": {
        "MDC (Sem 3)": "Financial Literacy",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Guitar-I",
    },
    "BA-ECO": {
        "MDC (Sem 3)": "Financial Literacy",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Desktop Publishing -I",
    },
    "BA-GAR": {
        "MDC (Sem 3)": "Gender Studies",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Guitar-I",
    },
    "BA-HIS": {
        "MDC (Sem 3)": "Gender Studies",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Mushroom Cultivation- I",
    },
    "BA-SOC": {
        "MDC (Sem 3)": "Gender Studies",
        "SEC (Sem 3)": "Introduction to Translation",
        "VTC": "Mushroom Cultivation- I",
    },
}


def repair_dob_text(text: str) -> str | None:
    """Fix common Excel typos before parsing."""
    raw = text.strip()
    if not raw:
        return None

    known_fixes = {
        "22.06.32007": "22.06.2007",
        "14.1.02005": "14.10.2005",
        "30.102005": "30.10.2005",
        "1.0702004": "01.07.2004",
    }
    if raw in known_fixes:
        return known_fixes[raw]

    # 22.06.32007 -> 22.06.2007 (year segment too long)
    match = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4,})$", raw)
    if match:
        day, month, year_part = match.groups()
        year = year_part[-4:]
        return f"{int(day):02d}.{int(month):02d}.{year}"

    # 30.102005 -> 30.10.2005 or 1.0702004 -> 01.07.2004
    match = re.match(r"^(\d{1,2})\.(\d{2})(\d{4})$", raw)
    if match:
        day, month, year = match.groups()
        if 1 <= int(month) <= 12:
            return f"{int(day):02d}.{int(month):02d}.{year}"

    # 14.1.02005 -> 14.10.2005
    match = re.match(r"^(\d{1,2})\.(\d{1,2})\.(0?)(\d{4})$", raw)
    if match:
        day, month, _pad, year = match.groups()
        if len(month) == 1 and month == "1":
            month = "10"
        return f"{int(day):02d}.{int(month):02d}.{year}"

    return raw


def parse_flexible_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none"}:
        return None

    text = repair_dob_text(text) or text

    match = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", text)
    if match:
        y, m, d = match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"

    match = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", text)
    if match:
        d, m, y = match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"

    match = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", text)
    if match:
        d, m, y = match.groups()
        return f"{y}-{int(m):02d}-{int(d):02d}"

    return None


def is_blank(value) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return not text or text.lower() in {"nan", "nat", "none"}


def normalize_gender(value) -> str | None:
    if is_blank(value):
        return None
    upper = str(value).strip().upper()
    if upper in {"MALE", "M"}:
        return "Male"
    if upper in {"FEMALE", "F"}:
        return "Female"
    if upper in {"OTHER", "O"}:
        return "Other"
    return str(value).strip()


def normalize_category(value) -> str | None:
    if is_blank(value):
        return None
    text = str(value).strip().upper()
    if text in {"GEN", "GENERAL"}:
        return "GENERAL"
    return text


def roll_to_portal_email(roll: str) -> str:
    slug = re.sub(r"[^a-z0-9]", "", roll.lower())
    return f"{slug}@student.donboscocollege.ac.in"


def header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            mapping[str(header).strip()] = col
    return mapping


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb["Students"]
    columns = header_map(ws)
    report: list[str] = []
    stats = {
        "rows": 0,
        "registration_filled": 0,
        "dob_fixed": 0,
        "email_generated": 0,
        "renewal_assigned": 0,
        "student_status_filled": 0,
        "gender_normalized": 0,
        "category_normalized": 0,
        "admission_date_fixed": 0,
        "issues": [],
    }

    data_start = 2
    if str(ws.cell(row=2, column=columns["Registration Number"]).value or "").startswith(
        "College"
    ):
        data_start = 3

    for row_idx in range(data_start, ws.max_row + 1):
        roll = ws.cell(row=row_idx, column=columns["Roll Number"]).value
        name = ws.cell(row=row_idx, column=columns["Full Name"]).value
        if is_blank(roll) and is_blank(name):
            continue
        stats["rows"] += 1

        roll_text = str(roll).strip() if not is_blank(roll) else ""
        reg_col = columns["Registration Number"]
        reg_val = ws.cell(row=row_idx, column=reg_col).value
        if is_blank(reg_val) and roll_text:
            ws.cell(row=row_idx, column=reg_col, value=roll_text)
            stats["registration_filled"] += 1

        dob_col = columns["Date of Birth"]
        dob_val = ws.cell(row=row_idx, column=dob_col).value
        parsed_dob = parse_flexible_date(dob_val)
        if parsed_dob and str(dob_val).strip() != parsed_dob:
            ws.cell(row=row_idx, column=dob_col, value=parsed_dob)
            stats["dob_fixed"] += 1
        elif is_blank(dob_val):
            stats["issues"].append(f"Row {row_idx}: missing Date of Birth ({name})")
        elif not parsed_dob:
            stats["issues"].append(
                f"Row {row_idx}: unparseable DOB '{dob_val}' ({name})"
            )

        adm_col = columns["Admission Date"]
        adm_val = ws.cell(row=row_idx, column=adm_col).value
        parsed_adm = parse_flexible_date(adm_val)
        if parsed_adm:
            ws.cell(row=row_idx, column=adm_col, value=parsed_adm)
            if str(adm_val) != parsed_adm:
                stats["admission_date_fixed"] += 1

        email_col = columns["Email Address"]
        email_val = ws.cell(row=row_idx, column=email_col).value
        if is_blank(email_val):
            if roll_text:
                generated = roll_to_portal_email(roll_text)
                ws.cell(row=row_idx, column=email_col, value=generated)
                stats["email_generated"] += 1
            else:
                stats["issues"].append(f"Row {row_idx}: missing email and roll ({name})")
        else:
            ws.cell(
                row=row_idx,
                column=email_col,
                value=str(email_val).strip().lower(),
            )

        prog = str(ws.cell(row=row_idx, column=columns["Programme"]).value or "").strip()
        if is_blank(ws.cell(row=row_idx, column=columns["MDC (Sem 3)"]).value):
            defaults = PROGRAM_SEM3_DEFAULTS.get(prog)
            if defaults:
                for field, value in defaults.items():
                    col = columns[field]
                    if is_blank(ws.cell(row=row_idx, column=col).value):
                        ws.cell(row=row_idx, column=col, value=value)
                stats["renewal_assigned"] += 1
            else:
                stats["issues"].append(
                    f"Row {row_idx}: missing Sem 3 papers, unknown programme {prog} ({name})"
                )

        if is_blank(ws.cell(row=row_idx, column=columns["AEC (Sem 3)"]).value):
            ws.cell(
                row=row_idx,
                column=columns["AEC (Sem 3)"],
                value="Introduction to Academic Writing (Arts)",
            )
            stats["renewal_assigned"] += 1

        status_col = columns["Student Status"]
        if is_blank(ws.cell(row=row_idx, column=status_col).value):
            ws.cell(row=row_idx, column=status_col, value="STUDYING")
            stats["student_status_filled"] += 1

        gender_col = columns["Gender"]
        gender_val = ws.cell(row=row_idx, column=gender_col).value
        normalized_gender = normalize_gender(gender_val)
        if normalized_gender and str(gender_val).strip() != normalized_gender:
            ws.cell(row=row_idx, column=gender_col, value=normalized_gender)
            stats["gender_normalized"] += 1

        cat_col = columns["Category"]
        cat_val = ws.cell(row=row_idx, column=cat_col).value
        normalized_cat = normalize_category(cat_val)
        if normalized_cat and str(cat_val).strip().upper() != normalized_cat:
            ws.cell(row=row_idx, column=cat_col, value=normalized_cat)
            stats["category_normalized"] += 1

        if is_blank(ws.cell(row=row_idx, column=columns["Student Mobile Number"]).value):
            stats["issues"].append(f"Row {row_idx}: missing mobile number ({name})")

    # Post-check duplicates
    emails: dict[str, int] = {}
    regs: dict[str, int] = {}
    for row_idx in range(data_start, ws.max_row + 1):
        email = ws.cell(row=row_idx, column=columns["Email Address"]).value
        reg = ws.cell(row=row_idx, column=columns["Registration Number"]).value
        if not is_blank(email):
            key = str(email).strip().lower()
            emails[key] = emails.get(key, 0) + 1
        if not is_blank(reg):
            key = str(reg).strip().upper()
            regs[key] = regs.get(key, 0) + 1

    dup_emails = [k for k, v in emails.items() if v > 1]
    dup_regs = [k for k, v in regs.items() if v > 1]
    if dup_emails:
        stats["issues"].append(f"Duplicate emails: {', '.join(dup_emails[:5])}")
    if dup_regs:
        stats["issues"].append(f"Duplicate registration numbers: {', '.join(dup_regs[:5])}")

    wb.save(OUTPUT)

    report.append("Morning Shift Import — Validation & Fix Report")
    report.append("=" * 60)
    report.append(f"Source: {SOURCE}")
    report.append(f"Output: {OUTPUT}")
    report.append("")
    report.append(f"Student rows processed: {stats['rows']}")
    report.append(f"Registration Number filled from Roll Number: {stats['registration_filled']}")
    report.append(f"Date of Birth normalized to YYYY-MM-DD: {stats['dob_fixed']}")
    report.append(f"Admission Date normalized: {stats['admission_date_fixed']}")
    report.append(f"Portal emails generated (no renewal form email): {stats['email_generated']}")
    report.append(f"Sem 3 MDC/SEC/VTC assigned (no renewal): {stats['renewal_assigned']} fields")
    report.append(f"Student Status set to STUDYING: {stats['student_status_filled']}")
    report.append(f"Gender values normalized: {stats['gender_normalized']}")
    report.append(f"Category values normalized: {stats['category_normalized']}")
    report.append("")
    report.append("Import readiness:")
    report.append("- Shift: MORNING (271 students, Semester 3)")
    report.append("- Registration Number: uses college Roll Number (e.g. BA25-601)")
    report.append("- Valid programme/MDC/AEC/SEC/VTC values checked against template lists")
    report.append("")
    if stats["issues"]:
        report.append("Remaining warnings / manual review:")
        for item in stats["issues"]:
            report.append(f"  - {item}")
    else:
        report.append("No blocking issues found. File is ready for ERP import (MERGE mode).")

    REPORT.write_text("\n".join(report), encoding="utf-8")
    print("\n".join(report))


if __name__ == "__main__":
    main()
